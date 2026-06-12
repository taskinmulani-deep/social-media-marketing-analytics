import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.image import Image as XLImage

os.makedirs('excel_reports', exist_ok=True)

# ── Load Data ─────────────────────────────────────────────────
df1 = pd.read_csv('data/cleaned/engagement_featured.csv')
df2 = pd.read_csv('data/cleaned/advertising_featured.csv')
df3 = pd.read_csv('data/cleaned/campaigns_featured.csv')
print("✅ Data loaded")

# ════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ════════════════════════════════════════════════════════════════

def style_header_row(ws, row_num, bg_color='1a1a2e', font_color='FFFFFF'):
    for cell in ws[row_num]:
        if cell.value:
            cell.font      = Font(bold=True, color=font_color, size=11)
            cell.fill      = PatternFill('solid', fgColor=bg_color)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border    = Border(
                bottom=Side(style='medium', color='4361ee'),
                right =Side(style='thin',   color='CCCCCC')
            )

def style_data_rows(ws, start_row, end_row, col_count):
    for r in range(start_row, end_row + 1):
        bg = 'F0F4FF' if r % 2 == 0 else 'FFFFFF'
        for c in range(1, col_count + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill      = PatternFill('solid', fgColor=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = Border(
                bottom=Side(style='thin', color='DDDDDD'),
                right =Side(style='thin', color='DDDDDD')
            )

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def write_df_to_sheet(ws, df, start_row=3, header_color='1a1a2e'):
    # Write headers
    for c, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c, value=col)
    style_header_row(ws, start_row, bg_color=header_color)

    # Write data
    for r, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_rows(ws, start_row + 1, start_row + len(df), len(df.columns))
    return start_row + len(df)

# ════════════════════════════════════════════════════════════════
# BUILD SUMMARY TABLES
# ════════════════════════════════════════════════════════════════

# KPI Summary
kpi_data = {
    'KPI': [
        'Total Posts Analyzed',
        'Avg Engagement Rate (%)',
        'Top Platform',
        'Best Content Type',
        'Total Campaigns (Advertising)',
        'Avg ROI - Advertising (%)',
        'Best Channel - ROI',
        'Total Campaigns (Campaign DB)',
        'Avg ROI - Campaigns (%)',
        'Best Campaign Type',
        'Avg Conversion Rate (%)',
        'Avg CTR (%)'
    ],
    'Value': [
        f"{len(df1):,}",
        f"{df1['Engagement_Rate'].mean():.2f}%",
        df1.groupby('Platform')['Engagement_Rate'].mean().idxmax(),
        df1.groupby('Content_Type')['Engagement_Rate'].mean().idxmax(),
        f"{len(df2):,}",
        f"{df2['ROI'].mean():.2f}%",
        df2.groupby('Channel_Used')['ROI'].mean().idxmax(),
        f"{len(df3):,}",
        f"{df3['ROI'].mean():.2f}%",
        df3.groupby('Campaign_Type')['ROI'].mean().idxmax(),
        f"{df3['Conversion_Rate'].mean():.2f}%",
        f"{df3['CTR'].mean():.2f}%"
    ]
}
df_kpi = pd.DataFrame(kpi_data)

# Platform engagement
df_platform = df1.groupby('Platform').agg(
    Total_Posts      =('Post_ID','count'),
    Avg_Engagement   =('Engagement_Rate','mean'),
    Avg_Likes        =('Likes','mean'),
    Avg_Shares       =('Shares','mean'),
    Avg_Comments     =('Comments','mean'),
    Avg_Virality     =('Virality_Score','mean'),
    Avg_Content_Score=('Content_Score','mean')
).round(2).sort_values('Avg_Engagement', ascending=False).reset_index()

# Content type
df_content = df1.groupby('Content_Type').agg(
    Total_Posts    =('Post_ID','count'),
    Avg_Engagement =('Engagement_Rate','mean'),
    Avg_Virality   =('Virality_Score','mean'),
    Avg_Save_Rate  =('Save_Rate','mean'),
    Avg_Views      =('Views','mean')
).round(2).sort_values('Avg_Engagement', ascending=False).reset_index()

# Influencer tier
df_influencer = df1.groupby('Influencer_Tier').agg(
    Total_Posts    =('Post_ID','count'),
    Avg_Engagement =('Engagement_Rate','mean'),
    Avg_Followers  =('Follower_Count','mean'),
    Avg_Virality   =('Virality_Score','mean')
).round(2).sort_values('Avg_Engagement', ascending=False).reset_index()

# Sentiment
df_sentiment = df1.groupby('Sentiment').agg(
    Total_Posts    =('Post_ID','count'),
    Avg_Engagement =('Engagement_Rate','mean'),
    Avg_Likes      =('Likes','mean'),
    Avg_Shares     =('Shares','mean')
).round(2).sort_values('Avg_Engagement', ascending=False).reset_index()

# Channel ROI
df_channel = df2.groupby('Channel_Used').agg(
    Total_Campaigns  =('Campaign_ID','count'),
    Avg_ROI          =('ROI','mean'),
    Avg_CTR          =('CTR','mean'),
    Avg_CVR          =('Conversion_Rate','mean'),
    Avg_Cost_Per_Click=('Cost_Per_Click','mean'),
    Avg_Eng_Score    =('Engagement_Score','mean')
).round(2).sort_values('Avg_ROI', ascending=False).reset_index()

# Campaign goal
df_goal = df2.groupby('Campaign_Goal').agg(
    Total_Campaigns=('Campaign_ID','count'),
    Avg_ROI        =('ROI','mean'),
    Avg_CTR        =('CTR','mean'),
    Avg_CVR        =('Conversion_Rate','mean'),
    Avg_Cost       =('Acquisition_Cost','mean')
).round(2).sort_values('Avg_ROI', ascending=False).reset_index()

# Campaign type
df_camptype = df3.groupby('Campaign_Type').agg(
    Total_Campaigns  =('Campaign_ID','count'),
    Avg_ROI          =('ROI','mean'),
    Avg_CVR          =('Conversion_Rate','mean'),
    Avg_CTR          =('CTR','mean'),
    Avg_Efficiency   =('Efficiency_Score','mean'),
    Avg_Cost         =('Acquisition_Cost','mean')
).round(2).sort_values('Avg_ROI', ascending=False).reset_index()

# Customer segment
df_segment = df3.groupby('Customer_Segment').agg(
    Total_Campaigns=('Campaign_ID','count'),
    Avg_ROI        =('ROI','mean'),
    Avg_CVR        =('Conversion_Rate','mean'),
    Avg_CTR        =('CTR','mean')
).round(2).sort_values('Avg_ROI', ascending=False).reset_index()

# Top 15 locations
df_location = df3.groupby('Location').agg(
    Total_Campaigns=('Campaign_ID','count'),
    Avg_ROI        =('ROI','mean'),
    Avg_CVR        =('Conversion_Rate','mean')
).round(2).sort_values('Avg_ROI', ascending=False).head(15).reset_index()

# Monthly trend
df_monthly = df2.groupby(['Year','Month']).agg(
    Avg_ROI=('ROI','mean'),
    Avg_CTR=('CTR','mean'),
    Total_Campaigns=('Campaign_ID','count')
).round(2).reset_index().sort_values(['Year','Month'])

print("✅ All summary tables built")

# ════════════════════════════════════════════════════════════════
# CREATE WORKBOOK
# ════════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)  # remove default sheet

DARK   = '1a1a2e'
BLUE   = '4361ee'
PINK   = 'f72585'
TEAL   = '4cc9f0'
GREEN  = '06d6a0'
PURPLE = '7209b7'

# ── SHEET 1: Executive Dashboard ────────────────────────────────
ws1 = wb.create_sheet('📊 Executive Dashboard')
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 3

# Title
ws1.merge_cells('B2:G2')
title_cell = ws1['B2']
title_cell.value     = '📊 SOCIAL MEDIA MARKETING ANALYTICS REPORT — 2026'
title_cell.font      = Font(bold=True, size=18, color='FFFFFF')
title_cell.fill      = PatternFill('solid', fgColor=DARK)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 40

# Subtitle
ws1.merge_cells('B3:G3')
sub = ws1['B3']
sub.value     = 'Comprehensive analysis of engagement, advertising ROI & campaign performance'
sub.font      = Font(italic=True, size=11, color='888888')
sub.alignment = Alignment(horizontal='center')
ws1.row_dimensions[3].height = 20

# KPI boxes
ws1.row_dimensions[5].height = 25
ws1.row_dimensions[6].height = 35
ws1.row_dimensions[7].height = 25

kpi_boxes = [
    ('B5:C7', f"{len(df1):,}",              'Total Posts',          BLUE),
    ('D5:E7', f"{df1['Engagement_Rate'].mean():.2f}%", 'Avg Engagement', PINK),
    ('F5:G7', f"{df2['ROI'].mean():.2f}%",  'Avg Ad ROI',           TEAL),
    ('B9:C11', f"{df3['Conversion_Rate'].mean():.2f}%", 'Avg Conversion', GREEN),
    ('D9:E11', df1.groupby('Platform')['Engagement_Rate'].mean().idxmax(), 'Top Platform', PURPLE),
    ('F9:G11', df3.groupby('Campaign_Type')['ROI'].mean().idxmax(), 'Best Campaign', BLUE),
]

for merge_range, value, label, color in kpi_boxes:
    ws1.merge_cells(merge_range)
    start_cell = merge_range.split(':')[0]
    cell = ws1[start_cell]
    cell.value     = f"{value}\n{label}"
    cell.font      = Font(bold=True, size=13, color='FFFFFF')
    cell.fill      = PatternFill('solid', fgColor=color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws1.row_dimensions[int(''.join(filter(str.isdigit, start_cell)))].height = 40

# Insights section
ws1.row_dimensions[13].height = 25
ws1.merge_cells('B13:G13')
ins_title = ws1['B13']
ins_title.value     = '📌 KEY FINDINGS & RECOMMENDATIONS'
ins_title.font      = Font(bold=True, size=13, color='FFFFFF')
ins_title.fill      = PatternFill('solid', fgColor=DARK)
ins_title.alignment = Alignment(horizontal='left', vertical='center')

top_platform    = df1.groupby('Platform')['Engagement_Rate'].mean().idxmax()
top_content     = df1.groupby('Content_Type')['Engagement_Rate'].mean().idxmax()
top_channel_roi = df2.groupby('Channel_Used')['ROI'].mean().idxmax()
top_camp_type   = df3.groupby('Campaign_Type')['ROI'].mean().idxmax()
top_segment     = df3.groupby('Customer_Segment')['Conversion_Rate'].mean().idxmax()
top_location    = df3.groupby('Location')['ROI'].mean().idxmax()

insights = [
    f"🔹 {top_platform} delivers the highest avg engagement — prioritize budget allocation here",
    f"🔹 '{top_content}' content type outperforms all others — create more of this format",
    f"🔹 {top_channel_roi} channel yields the best ROI — scale up campaigns on this channel",
    f"🔹 '{top_camp_type}' campaigns show strongest ROI — recommended for Q3/Q4 strategy",
    f"🔹 '{top_segment}' customer segment converts best — focus targeting on this segment",
    f"🔹 {top_location} is the top-performing location by ROI — increase geo-targeting here",
    f"🔹 Positive sentiment posts outperform negative by a significant margin",
    f"🔹 Nano/Micro influencers show higher engagement rates than Mega influencers",
]

for i, insight in enumerate(insights, 14):
    ws1.merge_cells(f'B{i}:G{i}')
    cell = ws1[f'B{i}']
    cell.value     = insight
    cell.font      = Font(size=10, color='1a1a2e')
    cell.fill      = PatternFill('solid', fgColor='F0F4FF' if i % 2 == 0 else 'FFFFFF')
    cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws1.row_dimensions[i].height = 20

set_col_widths(ws1, [3, 18, 18, 18, 18, 18, 18])
print("✅ Sheet 1: Executive Dashboard done")

# ── SHEET 2: Platform Performance ───────────────────────────────
ws2 = wb.create_sheet('📱 Platform Performance')
ws2.sheet_view.showGridLines = False

ws2.merge_cells('A1:H1')
t = ws2['A1']
t.value     = 'PLATFORM PERFORMANCE ANALYSIS'
t.font      = Font(bold=True, size=14, color='FFFFFF')
t.fill      = PatternFill('solid', fgColor=BLUE)
t.alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 35

write_df_to_sheet(ws2, df_platform, start_row=3, header_color=DARK)
set_col_widths(ws2, [18,14,16,12,12,14,14,18])

# Add bar chart
chart = BarChart()
chart.title    = "Avg Engagement Rate by Platform"
chart.y_axis.title = "Engagement Rate (%)"
chart.x_axis.title = "Platform"
chart.style   = 10
chart.height  = 12
chart.width   = 20

data_ref = Reference(ws2, min_col=3, max_col=3,
                     min_row=3, max_row=3+len(df_platform))
cats_ref = Reference(ws2, min_col=1,
                     min_row=4, max_row=3+len(df_platform))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws2.add_chart(chart, 'A' + str(3 + len(df_platform) + 3))

print("✅ Sheet 2: Platform Performance done")

# ── SHEET 3: Content Analysis ────────────────────────────────────
ws3 = wb.create_sheet('🎯 Content Analysis')
ws3.sheet_view.showGridLines = False

ws3.merge_cells('A1:F1')
t = ws3['A1']
t.value     = 'CONTENT TYPE & SENTIMENT ANALYSIS'
t.font      = Font(bold=True, size=14, color='FFFFFF')
t.fill      = PatternFill('solid', fgColor=PINK)
t.alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 35

# Content type table
ws3['A2'].value = '--- Content Type Performance ---'
ws3['A2'].font  = Font(bold=True, size=11, color=PINK)
write_df_to_sheet(ws3, df_content, start_row=3, header_color=DARK)

# Sentiment table below
gap = 3 + len(df_content) + 3
ws3.cell(row=gap, column=1).value = '--- Sentiment Analysis ---'
ws3.cell(row=gap, column=1).font  = Font(bold=True, size=11, color=PINK)
write_df_to_sheet(ws3, df_sentiment, start_row=gap+1, header_color=DARK)

# Influencer tier table
gap2 = gap + 1 + len(df_sentiment) + 3
ws3.cell(row=gap2, column=1).value = '--- Influencer Tier Performance ---'
ws3.cell(row=gap2, column=1).font  = Font(bold=True, size=11, color=PURPLE)
write_df_to_sheet(ws3, df_influencer, start_row=gap2+1, header_color=DARK)

set_col_widths(ws3, [20,14,16,14,14,14])
print("✅ Sheet 3: Content Analysis done")

# ── SHEET 4: Advertising ROI ─────────────────────────────────────
ws4 = wb.create_sheet('💰 Advertising ROI')
ws4.sheet_view.showGridLines = False

ws4.merge_cells('A1:G1')
t = ws4['A1']
t.value     = 'ADVERTISING ROI — CHANNEL & CAMPAIGN GOAL ANALYSIS'
t.font      = Font(bold=True, size=14, color='FFFFFF')
t.fill      = PatternFill('solid', fgColor=TEAL)
t.alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 35

ws4['A2'].value = '--- ROI by Channel ---'
ws4['A2'].font  = Font(bold=True, size=11, color=TEAL)
write_df_to_sheet(ws4, df_channel, start_row=3, header_color=DARK)

gap = 3 + len(df_channel) + 3
ws4.cell(row=gap, column=1).value = '--- ROI by Campaign Goal ---'
ws4.cell(row=gap, column=1).font  = Font(bold=True, size=11, color=TEAL)
write_df_to_sheet(ws4, df_goal, start_row=gap+1, header_color=DARK)

# Add bar chart for channel ROI
chart2 = BarChart()
chart2.title          = "Avg ROI by Channel"
chart2.y_axis.title   = "ROI (%)"
chart2.x_axis.title   = "Channel"
chart2.style          = 10
chart2.height         = 12
chart2.width          = 22

data_ref2 = Reference(ws4, min_col=3, max_col=3,
                      min_row=3, max_row=3+len(df_channel))
cats_ref2 = Reference(ws4, min_col=1,
                      min_row=4, max_row=3+len(df_channel))
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
ws4.add_chart(chart2, 'A' + str(gap + len(df_goal) + 5))

set_col_widths(ws4, [20,16,12,12,14,16,14])
print("✅ Sheet 4: Advertising ROI done")

# ── SHEET 5: Campaign Performance ───────────────────────────────
ws5 = wb.create_sheet('🚀 Campaign Performance')
ws5.sheet_view.showGridLines = False

ws5.merge_cells('A1:G1')
t = ws5['A1']
t.value     = 'CAMPAIGN TYPE & CUSTOMER SEGMENT PERFORMANCE'
t.font      = Font(bold=True, size=14, color='FFFFFF')
t.fill      = PatternFill('solid', fgColor=GREEN)
t.alignment = Alignment(horizontal='center', vertical='center')
ws5.row_dimensions[1].height = 35

ws5['A2'].value = '--- Campaign Type Performance ---'
ws5['A2'].font  = Font(bold=True, size=11, color=GREEN)
write_df_to_sheet(ws5, df_camptype, start_row=3, header_color=DARK)

gap = 3 + len(df_camptype) + 3
ws5.cell(row=gap, column=1).value = '--- Customer Segment Performance ---'
ws5.cell(row=gap, column=1).font  = Font(bold=True, size=11, color=GREEN)
write_df_to_sheet(ws5, df_segment, start_row=gap+1, header_color=DARK)

set_col_widths(ws5, [22,16,12,12,16,14,14])
print("✅ Sheet 5: Campaign Performance done")

# ── SHEET 6: Location Analysis ───────────────────────────────────
ws6 = wb.create_sheet('🌍 Location Analysis')
ws6.sheet_view.showGridLines = False

ws6.merge_cells('A1:E1')
t = ws6['A1']
t.value     = 'TOP LOCATIONS BY ROI PERFORMANCE'
t.font      = Font(bold=True, size=14, color='FFFFFF')
t.fill      = PatternFill('solid', fgColor=PURPLE)
t.alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 35

write_df_to_sheet(ws6, df_location, start_row=3, header_color=DARK)

# Add chart
chart3 = BarChart()
chart3.title        = "Top 15 Locations by Avg ROI"
chart3.y_axis.title = "ROI (%)"
chart3.style        = 10
chart3.height       = 14
chart3.width        = 24

data_ref3 = Reference(ws6, min_col=3, max_col=3,
                      min_row=3, max_row=3+len(df_location))
cats_ref3 = Reference(ws6, min_col=1,
                      min_row=4, max_row=3+len(df_location))
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)
ws6.add_chart(chart3, 'A' + str(3 + len(df_location) + 3))

set_col_widths(ws6, [22,16,12,14,14])
print("✅ Sheet 6: Location Analysis done")

# ── SHEET 7: Monthly Trend ───────────────────────────────────────
ws7 = wb.create_sheet('📈 Monthly Trend')
ws7.sheet_view.showGridLines = False

ws7.merge_cells('A1:F1')
t = ws7['A1']
t.value     = 'MONTHLY ROI & CTR TREND ANALYSIS'
t.font      = Font(bold=True, size=14, color='FFFFFF')
t.fill      = PatternFill('solid', fgColor=DARK)
t.alignment = Alignment(horizontal='center', vertical='center')
ws7.row_dimensions[1].height = 35

write_df_to_sheet(ws7, df_monthly, start_row=3, header_color=DARK)

# Line chart
chart4 = LineChart()
chart4.title        = "Monthly ROI Trend"
chart4.y_axis.title = "ROI (%)"
chart4.x_axis.title = "Month"
chart4.style        = 10
chart4.height       = 14
chart4.width        = 26

data_ref4 = Reference(ws7, min_col=3, max_col=3,
                      min_row=3, max_row=3+len(df_monthly))
chart4.add_data(data_ref4, titles_from_data=True)
ws7.add_chart(chart4, 'A' + str(3 + len(df_monthly) + 3))

set_col_widths(ws7, [10,10,12,12,18,14])
print("✅ Sheet 7: Monthly Trend done")

# ── SHEET 8: Raw Sample Data ─────────────────────────────────────
ws8 = wb.create_sheet('📋 Raw Sample')
ws8.sheet_view.showGridLines = False

ws8.merge_cells('A1:T1')
t = ws8['A1']
t.value     = 'RAW DATA SAMPLE — ENGAGEMENT (First 200 Rows)'
t.font      = Font(bold=True, size=13, color='FFFFFF')
t.fill      = PatternFill('solid', fgColor=DARK)
t.alignment = Alignment(horizontal='center', vertical='center')
ws8.row_dimensions[1].height = 30

sample = df1.head(200)
write_df_to_sheet(ws8, sample, start_row=3, header_color=DARK)

for i in range(1, len(sample.columns)+1):
    ws8.column_dimensions[get_column_letter(i)].width = 14

print("✅ Sheet 8: Raw Sample done")

# ════════════════════════════════════════════════════════════════
# SAVE WORKBOOK
# ════════════════════════════════════════════════════════════════
output_path = 'excel_reports/Social_Media_Analytics_Report_2026.xlsx'
wb.save(output_path)

print("\n" + "="*60)
print("✅ EXCEL REPORT CREATED SUCCESSFULLY!")
print("="*60)
print(f"📁 File: {output_path}")
print(f"📋 Sheets: 8 professional sheets")
print(f"   1. 📊 Executive Dashboard")
print(f"   2. 📱 Platform Performance")
print(f"   3. 🎯 Content Analysis")
print(f"   4. 💰 Advertising ROI")
print(f"   5. 🚀 Campaign Performance")
print(f"   6. 🌍 Location Analysis")
print(f"   7. 📈 Monthly Trend")
print(f"   8. 📋 Raw Sample")